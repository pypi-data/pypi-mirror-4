"""Parsers and writes to convert DataFrames in various (text) file formats.

"""
import csv
import gzip
import zlib
import itertools
import math
import os
import numpy
import sqlite3
import codecs
import tempfile

from dataframe import DataFrame
from factors import Factor

sqlite3.register_adapter(numpy.int32, int)
sqlite3.register_adapter(numpy.int64, int)
sqlite3.register_adapter(numpy.float64, float)
sqlite3.register_adapter(numpy.ma.core.MaskedArray, lambda x: numpy.nan)

class ParserError(ValueError):
    pass


def _intify(na_text):
    """Return a function that converts x to int or nan if x == naText"""
    def convert(value):
        """inner converter"""
        if value == na_text:
            return numpy.nan
        else:
            return int(value)
    return convert


def _floatify(na_text):
    """Return a function that converts x to float or nan if x == naText"""
    def convert(value):
        """inner converter"""
        if value == na_text:
            return numpy.nan
        else:
            return float(value)
    return convert


def _unquote(quote_char, row):
    """Remove quote characters from all value in row if present."""
    for i in xrange(0, len(row)):
        if row[i] and row[i][0] == quote_char and row[i][-1] == quote_char:
            row[i] = row[i][1:-1]


def _unquote_single_value(quote_char, value):
    """Remove quote characters at the beginning and end if present."""
    if value[0] == quote_char and value[-1] == quote_char:
        return value[1:-1]
    else:
        return value


def _open_file(filename_or_file_object, mode = 'rb', encoding=None):
    """Smartly open a file.

        Takes either an open file (returning it),
        a filename without .gz (opens it)
        or a filename with .gz (opens it as gzip file).

    """
    if not hasattr(filename_or_file_object, 'readlines'):
        # inspired by R's read.csv
        if filename_or_file_object.endswith('.gz'):
            filehandle = gzip.GzipFile(filename_or_file_object, mode)
        else:
            if encoding:
                filehandle = codecs.open(filename_or_file_object, mode, encoding=encoding)
            else:
                filehandle = open(filename_or_file_object, mode)
    else:
        filehandle = filename_or_file_object
    return filehandle

def read(filename): 
    """Try to read a df from a file, guessing by file extension which parser to use"""
    if filename.endswith('.xls'):
        import xlrd
        try:
            df = DF2Excel.read(filename)
        except xlrd.XLRDError:
            df = DF2TSV().read(filename)
    elif filename.endswith('.tsv'): 
        df = DF2TSV().read(filename)
    elif filename.endswith('.csv'): 
        df = DF2CSV().read(filename)
    elif filename.endswith('.xlsx'):
        df = DF2XLSX().read(filename)
    else:
        raise ValueError("Don't know how to parse %s" % filename)
    return df

def write(df, outputfilename):
    """Decide upon file extension how to write this dataframe..."""
    if outputfilename.endswith('.tsv'):
        DF2CSV().write(df, outputfilename, dialect=TabDialect())
    elif outputfilename.endswith('.csv'):
        DF2CSV().write(df, outputfilename)
    elif outputfilename.endswith('.xls'):
        DF2Excel().write(df, outputfilename)
    elif filename.endswith('.xlsx'):
        DF2XLSX().write(df, outputfilename)
    else:
        raise ValueError("Don't know how to handle %s " % outputfilename)

def parse_md_string(md_string):
    """Parse an MD string like and return a list of positions where it mismatched the reference
    deletions raise an exception for now...
    '36' (36 straight matches
    '10A25' (10 matches, an A that is mismatched, 25 letters matched
    '24A4A6' (24 matches, an A, 4 more letters, a mismatched A, 6 more letters
    '10A5^AC6' (10 matches, an A, 5 matches, a deletion of AC, 6 more letters
    (how are insertiens represented?)
    
    Unfortunatly, there is no way to capture this with a single regexps (it can be validated
    with the one in the SAM format spec, but python only keeps the last group of nested groups...)
    So a true lexer it is...
    """
    import re
    tokens = re.findall("([0-9]+|[A-Z]+|^)", md_string)
    tokens.reverse()
    try:
        pos = int(tokens.pop())
    except:
        raise ValueError("%s was not a valid md string, did not start with a number." % md_string)
    res = []
    while tokens:
        h = tokens.pop()
        if h == '^':
            raise ValueError("parse_md_string currently does not handle deletions (and insertions)")
        elif ord('0') <= ord(h[0]) <= ord('9'):
            pos += int(h)
        else:
            for letter in h:
                res.append(pos)
                pos += 1
    return res



import cStringIO
import codecs
class UnicodeWriter:
    """
    A CSV writer which will write rows to CSV file "f",
    which is encoded in the given encoding.
    """

    def __init__(self, f, dialect=csv.excel, encoding="utf-8", **kwds):
        # Redirect output to a queue
        self.queue = cStringIO.StringIO()
        self.writer = csv.writer(self.queue, dialect=dialect, **kwds)
        self.stream = f
        self.encoder = codecs.getincrementalencoder(encoding)()

    def writerow(self, row):
        self.writer.writerow([unicode(s).encode("utf-8") for s in row])
        # Fetch UTF-8 output from the queue ...
        data = self.queue.getvalue()
        data = data.decode("utf-8")
        # ... and reencode it into the target encoding
        data = self.encoder.encode(data)
        # write to the target stream
        self.stream.write(data)
        # empty queue
        self.queue.truncate(0)

    def writerows(self, rows):
        for row in rows:
            self.writerow(row)

class DF2CSV:
    """A converter for comma-seperated-value files and DataFrames."""

    def read(self, filename_or_file_object, header=True, dialect=None, 
        skip_rows = 0, columns_to_include = None, columns_to_exclude = None, 
        na_text = "NA", type_hints = None, handle_quotes = False, 
        has_row_names = False, r_skipped_first_column_name = False, 
        comment_character = None, encoding='latin-1', rename_duplicate_columns = False):
        """Read a csv file and turn it into a DataFrame.
        
        ::
        
            dialect = a dialect (object) valid for csv.reader
            skipRows = skip n rows at the start of the file
            columnsToInclude = column names if you want just a few
            columnsToExeclude = column names if you want to exclude some
            naText = entries matching will be converted to NaN
            typeHints = {columnName: numpy.dtype} if you want to enforce
                some types a priori.
            handleQuotes = if your csv is quoted (outside of the header)
                    if it isn't it can be read faster.
            r_skipped_first_column_name = if you export from R with write.table(row.names=TRUE),
                r skips the first column name. This handles this.
            rename_duplicate_columns - if a column name occurs multiple times, the second one is labeled _2, and so on.

        """
        if isinstance(header, csv.Dialect):
            raise ValueError("You passed a cvs.Dialect as second parameter, but that needs to be a bool. Pass your dialect as third (or dialect=) parameter.")
        try:
            filehandle = _open_file(filename_or_file_object,'r', encoding=encoding) 
            return self.__encapsulated_read(filehandle,
                                           header, dialect, skip_rows, columns_to_include, 
                                           columns_to_exclude, na_text, type_hints, handle_quotes, 
                                           has_row_names, r_skipped_first_column_name, comment_character, rename_duplicate_columns)
        except (zlib.error, IOError), e:
            if isinstance(e, zlib.error) or str(e).find('CRC check failed') != -1:
                filehandle = _open_file(filename_or_file_object,'r') 
                return self.__encapsulated_read(filehandle,
                                               header, dialect, skip_rows, columns_to_include, 
                                               columns_to_exclude, na_text, type_hints, handle_quotes, 
                                               has_row_names, r_skipped_first_column_name, comment_character, rename_duplicate_columns)
            else:
                raise e

    def __encapsulated_read(self, filehandle, header=True, dialect=None, skip_rows = 0, 
        columns_to_include = None, columns_to_exclude = None, na_text = "NA", type_hints = None, 
        handle_quotes = False, has_row_names = False, r_skipped_first_column_name = False, 
        comment_character = None, rename_duplicate_columns = False):
        if dialect is None:
            dialect = ExcelDialect()
        if isinstance(dialect, str):
            dialect = csv.get_dialect(dialect)

        if skip_rows: #skip's the first few rows
            for i in xrange(0, skip_rows):
                filehandle.readline()
        if handle_quotes:
            reader = csv.reader(filehandle, dialect=dialect)
            split_lines = [ row for row in reader ] # newline problems sometimes happen here - de-Macify...
            if comment_character:
                while split_lines[0][0] and split_lines[0][0][0] == comment_character:
                    split_lines = split_lines[1:]
            first_line = split_lines[0]
            if not (header is None or header is False):
                split_lines.pop(0)
            columns = list(itertools.izip_longest(*split_lines, fillvalue=na_text))
        else:
            current_pos = filehandle.tell()
            first_line = filehandle.readline().strip().split(dialect.delimiter)
            if comment_character:
                while first_line[0].startswith(comment_character):
                    first_line = filehandle.readline().strip().split(dialect.delimiter)
            if header is None or header is False:
                filehandle.seek(current_pos, os.SEEK_SET)
            column_no = len(first_line)
            if r_skipped_first_column_name:
                column_no += 1
            columns = self._csvToColumns(filehandle, dialect, column_no, na_text, comment_character)
        filehandle.close()
        # check/convert values
        if header:
            _unquote(dialect.quotechar, first_line)
            if header is True:
                fields = first_line
            else:
                fields = header
            if r_skipped_first_column_name:
                fields.insert(0,"Row")
            if len(fields) != len(first_line):
                should = len(first_line)
                if r_skipped_first_column_name:
                    should += 1
                raise ValueError("Wrong number of headers. Should be %i, was %i (potentially including skipped first column)" % (should, len(fields)))
            if len(fields) != len(set(fields)):
                if rename_duplicate_columns:
                    new_fields = []
                    for f in fields:
                        counter = 2
                        org_f = f
                        while f in new_fields:
                            f = "%s_%i" % ( org_f, counter )
                            counter += 1
                        new_fields.append(f)
                    fields = new_fields
                else:
                    raise ValueError("Some fields occur twice")
        else:
            num_fields = len(first_line)
            num_chars = int(math.ceil(math.log10(num_fields)))
            name_string = "column%0"+str(num_chars)+"d"
            fields = [ name_string % (i) for i in range(len(first_line)) ]
       
        # check/convert values
        if type_hints is None:
            type_hints = {}
        for i in xrange(0, len(columns)):
            try:
                columns[i] = map(_intify(na_text), columns[i])
                if not fields[i] in type_hints:
                    type_hints[fields[i]] = numpy.int32
            except ValueError:
                try:
                    columns[i] = map(_floatify(na_text), columns[i]) 
                except ValueError:
                    if i >= len(fields):
                        fields.append('Column_%i' % i)
                    if not fields[i] in type_hints:#keep as strings
                        max_len = len(max(columns[i], key=len))
                        dtype = 'S%i' % max_len
                        type_hints[fields[i]] = numpy.dtype


        results = {}
        fields_used = []
        for i in range(len(fields)):
            field = fields[i]
            try:
                if (columns_to_include is None or field in columns_to_include) and (columns_to_exclude is None or (not field in columns_to_exclude)):
                    if columns_to_include is not None and field in columns_to_include:
                        # Only use the first instance of an included column
                        columns_to_include.remove(field)
                    fields_used.append(field)
                    if field in type_hints:
                        try:
                            results[field] = numpy.array(columns[i], type_hints[field])
                        except ValueError:
                            if type_hints[field] is numpy.int32:
                                results[field] = numpy.array(columns[i], numpy.float64)
                            else:
                                raise
                    elif isinstance(columns[i], numpy.ma.core.MaskedArray):
                        results[field] = (columns[i])
                    else:
                        results[field] = list(columns[i])
            except IndexError:
                if columns: #so you can import 'empty' dfs
                    raise ValueError("Could not find column for field %s, file: %s" % (field, filehandle.name) )
                else: 
                    results[field] = []
        row_names = None
        if has_row_names:
            if isinstance(has_row_names, str): 
                row_names = results[has_row_names]
                del results[has_row_names]
                fields_used.remove(has_row_names)
            else:
                row_names = results[fields_used[0]]
                del results[fields_used[0]]
                del fields_used[0]
        return DataFrame( results, columns_ordered=fields_used, row_names_ordered = row_names )

    def _truncate_if_necessary(self, value, dialect):
        if dialect.max_entry_size:
            if isinstance(value, str) and len(value) > dialect.max_entry_size:
                return value[:dialect.max_entry_size] + '...'
        return value
    
    def write(self, data_frame, filename_or_file_object, dialect='excel'):
        """Write a DataFrame to a comma seperated value file."""
        filehandle = _open_file(filename_or_file_object, 'wu')
        if dialect == 'excel':
            dialect = ExcelDialect()
        #writer = csv.writer(filehandle, dialect=dialect)
        writer = UnicodeWriter(filehandle, dialect=dialect)
        if not data_frame.row_names is None:
            header = ["Row"]
            header.extend(data_frame.columns_ordered)
        else:
            header = data_frame.columns_ordered 
        writer.writerow(header)
        if not data_frame.row_names is None:
            rowNames = data_frame.row_names
            for i in range(data_frame.num_rows):
                row = [rowNames[i]]
                row.extend(self._truncate_if_necessary(x, dialect) for x in data_frame.get_row_as_list(i))
                writer.writerow(row)
        else:
            for i in range(data_frame.num_rows):
                writer.writerow([self._truncate_if_necessary(x, dialect) for x in data_frame.get_row_as_list(i)])
        if not hasattr(filehandle, 'getvalue'): #this should cover both stringio and cStringIO - the later doesn't export a class, apparently.
            filehandle.close()


    def _csvToColumns(self, file_object, dialect, number_of_columns, na_text, comment_character):
        """Turn a csv file into a list of columns."""
        columns = []
        for col_no in xrange(0, number_of_columns):
            columns.append([])
        sep = dialect.delimiter
        for row in file_object:
            if comment_character and row.startswith(comment_character):
                continue
            row = row.rstrip().split(sep)
            #_unquote(dialect.quotechar, row)
            for i in xrange(0, number_of_columns):
                try:
                    columns[i].append(row[i])
                except IndexError:
                    columns[i].append(na_text)
        return columns

class DF2TSV(DF2CSV):
    """A L{DF2CSV()} with dialect permanently set to TabDialect"""

    def read(self, filename_or_file_object, **kwargs):
        kwargs['dialect'] = TabDialect()
        return DF2CSV.read(self,filename_or_file_object, **kwargs)

    def write(self, data_frame, filename_or_file_object, **kwargs):
        kwargs['dialect'] = TabDialect()
        return DF2CSV.write(self, data_frame, filename_or_file_object, **kwargs)




class DF2ARFF:
    """ARFF file format (Weka) to DataFrame converter"""

    def read(self, filename_or_file_object):
        """Read an ARFF file.

        """
        filehandle = _open_file(filename_or_file_object,'rb')
        attributes = []
        data_found = False
        rows = filehandle.readlines()
        filehandle.close()
        row_no = 0
        for row_no, row in enumerate(rows):
            if row.startswith('@attribute'):
                row = row.split()
                name = row[1]
                type_ = row[2]
                attributes.append((name, type_))
            elif row.startswith('@data'):
                data_found = True
                break
        if not data_found:
            raise ValueError("Not a valid arff file - no @data found")
        if not attributes:
            raise ValueError("Not a valid arff file - no @attribute found")
        data_start_row = row_no + 1
        num_rows = len(rows) - data_start_row 

        field_names = zip(*attributes)[0]
        fields = {}
        fields_in_order = []
        field_funcs_in_order = []
        for name, type_ in attributes:
            if type_ == 'numeric':
                fields[name] = numpy.ma.zeros((num_rows, ), dtype=numpy.double)
                field_funcs_in_order.append(float)
            elif type_.startswith('{') and type_.endswith('}'):
                values = type_[1:-1].split(",") #todo: this should really parse the values instead, they're quoted after all
                _unquote("'", values)
                max_len = len(max(values, key=len))
                fields[name] = numpy.ma.zeros((num_rows, ), dtype = 'S%i' % max_len)

                field_funcs_in_order.append(lambda x: _unquote_single_value("'", x))
            else:
                raise ValueError("Don't know yet how to handle type %s"  % type_)
            fields_in_order.append(fields[name])
        for row_no, row in enumerate(rows[data_start_row:]):
            row = row.strip().split(',')
            for i, value in enumerate(row):
                fields_in_order[i][row_no] = field_funcs_in_order[i](value)
        return DataFrame(fields, field_names)

    def write(self, data_frame, filename_or_file_object):
        """Write a DataFrame to an ARFF file for Weak.
        """
        raise NotImplementedError()

class DF2Excel:
    """Use xlrd and xlwt to write 'real' excel files"""

    def read(self, filename, sheet_name = None, row_offset = 0, col_offset = 0, row_max = None, col_max = None, filter_empty = False, NA_string = None, handle_encoding_errors = lambda u: u.encode('latin-1', 'replace').strip().decode('latin-1'), columns_to_include = None):
        import xlrd
        if filename.endswith('.gz'):
            tf = tempfile.NamedTemporaryFile()
            op = _open_file(filename,'rb')
            tf.write(op.read())
            op.close()
            tf.flush()
            filename = tf.name
        else:
            tf = None
        try:
            wb = xlrd.open_workbook(filename)
            if sheet_name is None:
                ws = wb.sheet_by_index(0)
            else:
                if type(sheet_name) == int:
                    ws = wb.sheet_by_index(sheet_name)
                else:
                    ws = wb.sheet_by_name(sheet_name)
            if row_max is None:
                row_max = ws.nrows
            if col_max is None:
                col_max = ws.ncols
            cols = {}
            col_names_in_order = []
            columns_used = []
            for yy, col_name in enumerate(ws.row(0 + row_offset)[col_offset: col_max]):
                name = col_name.value
                counter = 2
                while name in col_names_in_order:
                    name = "%s_%i" % ( col_name.value, counter )
                    counter += 1
                col_names_in_order.append(name)
                if not columns_to_include or name in columns_to_include:
                    cols[name] = []
                    columns_used.append(yy)
            if not columns_used:
                raise ValueError("No columns to import - columns_to_include did not reference any present. Column_to_include was %s, available: %s" % (columns_to_include, col_names_in_order))
            if columns_to_include and set(columns_to_include).difference(col_names_in_order):
                raise ValueError("Nonexistant column name(s) in columns_to_include: %s" % (set(columns_to_include).difference(col_names_in_order),))
            for row_no in xrange(0 + row_offset + 1, row_max):
                row = ws.row(row_no)
                found = False
                for y in  columns_used:#xrange(0 + col_offset, col_max):
                    value = row[y].value
                    try:
                        if str(value).strip() == '' or row[y].ctype == xlrd.XL_CELL_EMPTY or row[y].ctype == xlrd.XL_CELL_ERROR:
                            value = None
          #              cols[col_names_in_order[y - col_offset]].append(str(value))
                        if type(value) == unicode and value == NA_string:
                            value = None
                    except UnicodeEncodeError:
                        value = handle_encoding_errors(value)
                    cols[col_names_in_order[y - col_offset]].append((value))
                    if value:
                        found = True
                if not found and filter_empty:
                    for k in col_names_in_order:
                        cols[k].pop()
            return DataFrame(cols, [x for (yy,x) in enumerate(col_names_in_order) if yy in columns_used])
        finally:
            if tf:
                tf.close()


    def write(self, data_frame_or_dict_of_dataframes, filename_or_file_object, sheet_name = "DataFrame",
             highlight_columns = None):
        if isinstance(data_frame_or_dict_of_dataframes, DataFrame):
            data_frame_or_dict_of_dataframes = {sheet_name: data_frame_or_dict_of_dataframes}
        elif isinstance(data_frame_or_dict_of_dataframes, dict):
            pass
        else:
            raise ValueError("DF2Excel only writes out Dataframes, or dicts of {sheet_name: dataframe}")
        import xlwt
        filehandle = _open_file(filename_or_file_object, 'wb')
        wb = xlwt.Workbook()
        style_normal = xlwt.XFStyle() 
        style_highlight = style1 = xlwt.easyxf(""" 
             pattern: 
                 back_colour yellow, 
                 pattern solid, 
                 fore-colour lavender 
         """) 
        for sheet_name, data_frame in data_frame_or_dict_of_dataframes.items():
            if data_frame.num_rows >= 65535:
                raise ParserError("Too many rows, excel only supports 65k")
            ws = wb.add_sheet(sheet_name)
            i_col = 0
            i_row = 0
            #ws.write(i_col, i_row, "Row")
            #i_col += 1
            translate_value = self.translate_value
            for column_name in data_frame.columns_ordered:
                this_column = data_frame.get_column_view(column_name)
                if isinstance(this_column,Factor):
                    this_column = this_column.as_levels()
                if highlight_columns and column_name in highlight_columns:
                    style = style_highlight
                else:
                    style = style_normal
                ws.write(i_row, i_col, column_name, style)
                for i in range(0, data_frame.num_rows):
                    ws.write(i_row + i + 1, i_col, translate_value(this_column[i]), style)
                i_col += 1
                i_row = 0

        wb.save(filehandle)

    def translate_value(self, value):
        typ = type(value)
        if typ is str:
            if not value:
                return ""
            else:
                return unicode(value, 'latin-1', 'replace')[:65000]
        elif typ is float:
            if numpy.isnan(value):
                return None
            elif numpy.isinf(value):
                return "INF"
            else:
                return value
        elif typ is numpy.float64:
            if numpy.isnan(value):
                return None
            elif numpy.isinf(value):
                return "INF"
            else:
                return float(value)
        elif typ is int:
            return value
        elif typ is numpy.int32 or typ is numpy.int64:
            return int(value)
        elif typ is bool:
            return value
        elif value is None:
            return ""
        elif typ is numpy.ma.core.MaskedArray and value.mask:
            return ""
        else:
            try:
                return float(value)
            except:
                return unicode(value)[:65000]

class DF2XLSX:
    """Use openpyxl to read and xlxs excel files"""

    def read(self, filename, sheet_name = None, row_offset = 0, col_offset = 0, row_max = None, col_max = None, filter_empty = False, NA_string = None, handle_encoding_errors = lambda u: u.encode('latin-1', 'replace').strip().decode('latin-1')):
        import openpyxl
        wb = openpyxl.reader.excel.load_workbook(filename)
        if sheet_name is None:
            ws = wb.worksheets[0]
        else:
            if type(sheet_name) == int:
                ws = wb.worksheets[sheet_name]
            else:
                ws = wb.sheet_by_name(sheet_name)
        if row_max is None:
            row_max = len(ws.rows)
        if col_max is None:
            col_max = len(ws.columns)
        cols = {}
        col_names_in_order = []
        for col_name in ws.rows[0 + row_offset][col_offset: col_max]:
            name = col_name.value
            counter = 2
            while name in col_names_in_order:
                name = "%s_%i" % ( col_name.value, counter )
                counter += 1
            col_names_in_order.append(name)
            cols[name] = []
        for row_no in xrange(0 + row_offset + 1, row_max):
            row = ws.rows[row_no]
            found = False
            for y in xrange(0 + col_offset, col_max):
                value = row[y].value
                try:
                    if str(value).strip() == '': #or row[y].ctype == xlrd.XL_CELL_EMPTY or row[y].ctype == xlrd.XL_CELL_ERROR:
                        value = None
      #              cols[col_names_in_order[y - col_offset]].append(str(value))
                    if type(value) == unicode and value == NA_string:
                        value = None
                except UnicodeEncodeError:
                    value = handle_encoding_errors(value)
                cols[col_names_in_order[y - col_offset]].append((value))
                if value:
                    found = True
            if not found and filter_empty:
                for k in col_names_in_order:
                    cols[k].pop()
        return DataFrame(cols, col_names_in_order)

    def write(self, data_frame_or_dict_of_dataframes , filename_or_file_object, sheet_name = "DataFrame",
             highlight_columns = None):
        import openpyxl
        wb = openpyxl.workbook.Workbook()
        if not isinstance(data_frame_or_dict_of_dataframes, dict):
            data_frame_or_dict_of_dataframes = {sheet_name: data_frame_or_dict_of_dataframes}
        first = True
        for sheet_name, data_frame in data_frame_or_dict_of_dataframes.items():
            if first:
                ws = wb.get_active_sheet()
                first = False
            else:
                ws = wb.create_sheet()
            ws.title = sheet_name
            for xx, column_name in enumerate(data_frame.columns_ordered):
                ws.cell(row = 0, column = xx).value = column_name
                for yy, value in enumerate(data_frame.get_column_view(column_name)):
                    try:
                        ws.cell(row = yy + 1, column = xx).value = value
                    except TypeError:
                        ws.cell(row = yy + 1, column = xx).value = float(value)
        wb.save(filename_or_file_object)


class DF2Sqlite:

    def write(self, data_frame_or_dict_of_dataframes, filename_or_file_object, sheet_name = 'DataFrame'):
        if type(filename_or_file_object) is str and os.path.exists(filename_or_file_object):
            os.unlink(filename_or_file_object)
        conn = sqlite3.connect(filename_or_file_object,isolation_level="DEFERRED")
        cur = conn.cursor()
        if type(data_frame_or_dict_of_dataframes) is DataFrame:
            data_frame_or_dict_of_dataframes = {sheet_name: data_frame_or_dict_of_dataframes}
        cur.execute('BEGIN TRANSACTION')
        for table_name, data_frame in data_frame_or_dict_of_dataframes.items():
            #print self._create_table_statement(table_name, data_frame)
            cur.execute(self._create_table_statement(table_name, data_frame))
            ins_statement = "INSERT INTO '%s' VALUES (%s)" % (table_name, ",".join(["?"] * len(data_frame.columns_ordered)))
            cur.executemany(ins_statement, data_frame.iter_rows_as_list())
        conn.commit()
        conn.close()

    def map_df_values(self, data_frame):
        int_types = (numpy.int32, numpy.int64)
        float_types = (numpy.float64,)
        for row in data_frame.iter_rows_list():
            res = []
            for val in row:
                if type(val) in int_types:
                    val = int(val)
                elif type(val) in float_types:
                    val = float(val)
                elif type(val) is numpy.ma.core.MaskedArray:
                    val = numpy.nan
                res.append(val)
            yield res


    def _create_table_statement(self, table_name, data_frame):
        field_defs = []
        for column_name in data_frame.columns_ordered:
            dtype = data_frame.value_dict[column_name].dtype
            if dtype in (numpy.int32, numpy.uint32, numpy.int64, numpy.uint64, numpy.int8, numpy.uint8, numpy.int16, numpy.uint16):
                column_type = 'int'
            elif dtype in (numpy.float, numpy.float64):
                column_type = 'real'
            else:
                column_type = 'text'
            field_defs.append("'%s' %s" % (column_name, column_type))
        statement = 'CREATE TABLE \'%s\' (%s)' % (table_name, ",\n".join(field_defs)) 
        return statement

class DF2HTMLTable:

    def read(self, extracted_table):
        """A terribly simple table parser"""
        from BeautifulSoup import BeautifulSoup
        soup = BeautifulSoup(extracted_table)
        table = soup.table
        rows = table.findAll('tr')
        col_names = [] 
        for td in rows[0].findAll('td'):
            col_names.append(str(td.string))
        data = {}
        for name in col_names:
            data[name] = []
        for row in rows[1:]:
            for ii, td in enumerate(row.findAll('td')):
                if ii == len(col_names):
                    break
                data[col_names[ii]].append(str(td.string))
            for ii in xrange(ii + 1, len(col_names)):
                data[col_names[ii]].append('')
        return DataFrame(data)







class Access2000Dialect(csv.Dialect):
    """A dialect to properly interpret Microsoft Access2000 CSV exports 
    for international languages.
    """
    delimiter = ';'
    quotechar = '"'
    doublequote = True
    quoting = csv.QUOTE_NONNUMERIC
    lineterminator = '\n'
    skipinitialspace = True

class TabDialect(csv.Dialect):
    """A dialect to interpret tab separated files.
    """
    delimiter = '\t'
    quotechar = '"'
    doublequote = True
    quoting = csv.QUOTE_MINIMAL
    lineterminator = '\n'
    skipinitialspace = True
    max_entry_size = 65000 #entries are truncated during write and '...' is appended.

class CommaDialect(csv.Dialect):
    """A dialect to interpret tab separated files.
    """
    delimiter = ','
    quotechar = '"'
    doublequote = True
    quoting = csv.QUOTE_MINIMAL
    lineterminator = '\n'
    skipinitialspace = True
    max_entry_size = 65000
    
    def __init__(self, **args):
        for kw in args:
            setattr(self, kw, args[kw])


class SpaceDialect(csv.Dialect):
    """A dialect to interpret tab separated files.
    """
    delimiter = ' '
    quotechar = '"'
    doublequote = True
    quoting = csv.QUOTE_MINIMAL
    lineterminator = '\n'
    skipinitialspace = True
    max_entry_size = None

class ExcelDialect(csv.Dialect):
    delimiter = ';'
    quotechar = '"'
    doublequote = True
    quoting = csv.QUOTE_MINIMAL
    lineterminator = '\n'
    skipinitialspace = True
    max_entry_size = 65000

